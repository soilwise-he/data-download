# app/main.py
from fastapi import FastAPI, HTTPException, Body
from fastapi.responses import FileResponse, Response, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Any, Dict, Optional, Union
import httpx, json, io
import pandas as pd
from io import BytesIO, StringIO
import csv, os
from rdflib import Graph, Namespace, BNode, URIRef, Literal
from rdflib.namespace import RDF, RDFS, XSD, SKOS, DCTERMS
import sqlite3, sys
import shutil, hashlib, random
from datetime import datetime
from csvwlib import CSVWConverter

# Namespaces
SOSA = Namespace("http://www.w3.org/ns/sosa/")
GEO  = Namespace("http://www.opengis.net/ont/geosparql#")
WGS  = Namespace("http://www.w3.org/2003/01/geo/wgs84_pos#")
QUDT = Namespace("http://qudt.org/schema/qudt/")

rootpath = os.environ.get("ROOTPATH") or "/"


app = FastAPI(    
    title="Soil Observation Data Annotation with CSVW",
    description="""
API for converting soil observation data into RDF and SQLite formats.

### Features
- Convert CSVW to RDFor SQLITE
- Export full project
- Suggest a CSVW configuration
""", root_path=rootpath)

# DEV: permissive CORS for local testing. Lock this down in production.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

class URLRequest(BaseModel):
    url: str

class JsonLdPayload(BaseModel):
    MetadataContent: Any

class ConvertRequest(BaseModel):
    context: Union[dict, str]  # JSON object or URL string
    data: Optional[list[str]] = [] 
    output_format: Optional[str] = "ttl"        # 'ttl', 'rdfxml', 'json-ld'


# -----------------------
# Utilities
# -----------------------
async def fetch_bytes(url: str, timeout: int = 30) -> bytes:
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
        r = await client.get(url)
        r.raise_for_status()
        return r.content

def dataframe_to_rows_and_headers(df: pd.DataFrame) -> (List[str], List[Dict[str,Any]]):
    # ensure columns are strings
    headers = [str(c) for c in df.columns.tolist()]
    rows = df.fillna("").to_dict(orient="records")
    return headers, rows

def to_csv_text(headers: List[str], rows: List[Dict[str,Any]]) -> str:
    out = StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h, "") for h in headers])
    return out.getvalue()

# header detection for a sheet represented as list-of-lists
def detect_header_row(raw_rows: List[List[Any]], max_scan: int = 20) -> int:
    def is_empty_row(r):
        if not r:
            return True
        return all((c is None) or (str(c).strip() == "") for c in r)

    max_scan = min(max_scan, len(raw_rows))
    best_idx = 0
    best_score = -1e9

    for i in range(max_scan):
        row = raw_rows[i] or []
        if is_empty_row(row):
            continue
        non_empty = sum(1 for c in row if c is not None and str(c).strip() != "")
        if non_empty == 0:
            continue
        unique_count = len(set(str(c).strip() if c is not None else "" for c in row))
        uniqueness = unique_count / max(1, non_empty)

        # evaluate following rows as "data-like"
        following_good = 0
        following_checked = 0
        for j in range(i+1, min(len(raw_rows), i+1+10)):
            r2 = raw_rows[j] or []
            non_empty2 = sum(1 for c in r2 if c is not None and str(c).strip() != "")
            if non_empty2 == 0:
                following_checked += 1
                continue
            if non_empty2 >= max(1, int(non_empty * 0.5)):
                following_good += 1
            following_checked += 1

        follow_ratio = (following_good / following_checked) if following_checked > 0 else 0.0
        score = non_empty * 0.6 + uniqueness * 2.0 + follow_ratio * 5.0
        if uniqueness < 0.25:
            continue
        if score > best_score:
            best_score = score
            best_idx = i

    # fallback to first non-empty row
    if best_score < -1e8:
        for i in range(min(5, len(raw_rows))):
            if not is_empty_row(raw_rows[i]):
                return i
        return 0
    return best_idx

# guess type from samples (list of values)
def guess_type_from_samples(values: List[Any]) -> str:
    num_count = 0
    date_count = 0
    non_empty = 0
    for v in values:
        s = "" if v is None else str(v).strip()
        if s == "":
            continue
        non_empty += 1
        # numeric?
        try:
            ff = float(s.replace(",", "."))
            if pd.notna(ff):
                num_count += 1
                continue
        except Exception:
            pass
        # date-like?
        try:
            parsed = pd.to_datetime(s, errors="coerce")
            if pd.notna(parsed):
                date_count += 1
                continue
        except Exception:
            pass
    if non_empty == 0:
        return "string"
    if num_count / non_empty >= 0.8:
        return "number"
    if date_count / non_empty >= 0.8:
        return "date"
    return "string"

# pick primary key candidate from rows (list of dicts)
def pick_primary_key(headers: List[str], rows_sample: List[Dict[str,Any]]) -> Optional[str]:
    # compute unique ratio per header
    scores = []
    for h in headers:
        non_empty = 0
        seen = set()
        for r in rows_sample:
            v = r.get(h, "")
            s = "" if v is None else str(v).strip()
            if s == "":
                continue
            non_empty += 1
            seen.add(s)
        if non_empty == 0:
            uniq_ratio = 0.0
        else:
            uniq_ratio = len(seen) / non_empty
        scores.append((h, non_empty, len(seen), uniq_ratio))
    # prefer name hints
    hints = ['id', 'identifier', 'uuid', 'code', 'key']
    strong = [s for s in scores if s[1] >= 3 and s[3] >= 0.98]
    good = [s for s in scores if s[1] >= 3 and s[3] >= 0.8]
    def pick_by_hint(cands):
        for hint in hints:
            for c in cands:
                if hint in c[0].lower():
                    return c[0]
        return None
    if len(strong) == 1:
        return strong[0][0]
    if len(strong) > 1:
        return pick_by_hint(strong) or sorted(strong, key=lambda x: -x[3])[0][0]
    if good:
        return pick_by_hint(good) or sorted(good, key=lambda x: -x[3])[0][0]
    # final attempt: perfect uniqueness
    perfect = [s for s in scores if s[1] >= 3 and s[2] == s[1]]
    if perfect:
        return pick_by_hint(perfect) or perfect[0][0]
    return None

# -----------------------
# Core parsing logic
# -----------------------
async def parse_excel_or_csv_from_url(url: str) -> List[Dict]:
    """
    Returns a list of table entries:
      [{ "url": "<sheet-identifier-or-url>", "sheet_name": "<...>", "headers": [...], "rows": [...], "csv_text": "..." }, ...]
    """
    content = await fetch_bytes(url)
    # try to detect if excel by signature (PK header for xlsx) or by extension
    lower = url.lower()
    is_excel_ext = lower.endswith(".xlsx") or lower.endswith(".xls") or lower.endswith(".xlsm") or lower.endswith(".ods")
    # quick signature check: xlsx files are zip (PK\x03\x04)
    if (not is_excel_ext) and content[:4] != b'PK\x03\x04':
        # treat as CSV text
        text = content.decode("utf-8", errors="replace")
        parsed = list(csv.DictReader(StringIO(text)))
        headers = parsed[0].keys() if parsed else []
        headers = [str(h) for h in headers]
        rows = parsed
        csv_text = text
        return [{
            "url": url,
            "sheet_name": None,
            "headers": headers,
            "rows": rows,
            "csv_text": csv_text
        }]
    # else parse as excel using pandas
    bio = BytesIO(content)
    try:
        # sheet_name=None reads all sheets into dict
        xls = pd.read_excel(bio, sheet_name=None, engine="openpyxl")
    except Exception as e:
        # fallback: try using engine 'xlrd' for old xls (if available)
        raise HTTPException(status_code=500, detail=f"Failed to read Excel workbook: {e}")

    results = []
    # pandas returns dict of DataFrame
    for sheet_name, df in xls.items():
        # convert dataframe to raw rows (list-of-lists) to run header detection
        raw_rows = df.fillna(value=pd.NA).astype(object).values.tolist()
        # raw_rows may include header row as first row (pandas by default already uses first row as header)
        # But since we asked sheet_name=None with default behavior, pandas used first non-empty row as header.
        # To be conservative, convert sheet to array-of-arrays using openpyxl if you want raw cell values.
        # Here we use a fallback: if all column names are "Unnamed..." then detect header row from values
        col_names = df.columns.tolist()
        if all(str(c).startswith("Unnamed") for c in col_names):
            # try building raw_rows from values (pandas placed header as Unnamed)
            # Use df.to_records? We'll re-read via openpyxl for better raw data.
            try:
                from openpyxl import load_workbook
                bio.seek(0)
                wb = load_workbook(filename=bio, read_only=True, data_only=True)
                ws = wb[sheet_name]
                raw = []
                for row in ws.iter_rows(values_only=True):
                    raw.append(list(row))
                header_idx = detect_header_row(raw)
                header_row = ["" if c is None else str(c).strip() for c in raw[header_idx]]
                # build rows
                data_rows = []
                empty_streak = 0
                for r in raw[header_idx+1:]:
                    if r is None or all((c is None or str(c).strip()=="") for c in r):
                        empty_streak += 1
                        if empty_streak >= 5:
                            break
                        continue
                    empty_streak = 0
                    rowobj = {}
                    for j, h in enumerate(header_row):
                        key = h or f"col{j+1}"
                        val = r[j] if j < len(r) else ""
                        rowobj[key] = "" if val is None else val
                    data_rows.append(rowobj)
                headers = header_row
                csv_text = to_csv_text(headers, data_rows)
                results.append({
                    "url": f"{url}#{sheet_name}",
                    "sheet_name": sheet_name,
                    "headers": headers,
                    "rows": data_rows,
                    "csv_text": csv_text
                })
                continue
            except Exception:
                # fallback to using pandas parsing below
                pass

        # If we get here, use DataFrame as parsed by pandas.
        headers, rows = dataframe_to_rows_and_headers(df)
        csv_text = to_csv_text(headers, rows)
        results.append({
            "url": f"{url}#{sheet_name}",
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": rows,
            "csv_text": csv_text
        })
    return results

# Helper: safe filename from URL or fallback
def safe_name_from_url(url: str, fallback: str) -> str:
    try:
        parsed = urlparse(url)
        name = parsed.path.split("/")[-1] or parsed.netloc or fallback
        # remove fragments and query parts if present
        name = name.split("#")[0].split("?")[0]
        if not name:
            return fallback
        return name
    except Exception:
        return fallback

# Helper: build CSV text from header list + (optional) rows (rows as list of dicts)
def build_csv_text_from_rows(headers: List[str], rows: Optional[List[Dict[str, Any]]] = None) -> str:
    output = BytesIO()
    # write text through TextIOWrapper to support newline handling in zip
    text_wr = TextIOWrapper(output, encoding="utf-8", newline="")
    writer = csv.writer(text_wr, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    if rows:
        for r in rows:
            row = [ (r.get(h) if isinstance(r, dict) else '') for h in headers ]
            writer.writerow(row)
    text_wr.flush()
    result = output.getvalue().decode("utf-8")
    output.close()
    return result

# Async fetch helper with timeout and basic heuristics
async def fetch_text(client: httpx.AsyncClient, url: str, timeout: float = 10.0) -> Optional[str]:
    try:
        resp = await client.get(url, timeout=timeout)
        resp.raise_for_status()
        # try to decode as text
        ct = resp.headers.get("content-type", "")
        # treat as text if content-type mentions text or csv or json
        if "text" in ct or "csv" in ct or "json" in ct or ct == "" or "application/octet-stream" in ct:
            # prefer resp.text which decodes according to charset
            return resp.text
        # otherwise try to decode with utf-8 fallback
        return resp.content.decode("utf-8", errors="replace")
    except Exception:
        return None

@app.post("/export", summary="Exports a CSVW project (tables and configuration) as a zip",
    description="""Accepts JSON body with { "MetadataContent": <object|json-string> }
    Produces a ZIP with:
      - one CSV file per table (try fetch table.url, else generate from tableSchema.columns headers)
      - metadata.jsonld (the metadataPayload as JSON)""")
async def export(req: URLRequest):
    """
    Accepts JSON body with { "MetadataContent": <object|json-string> }
    Produces a ZIP with:
      - one CSV file per table (try fetch table.url, else generate from tableSchema.columns headers)
      - metadata.jsonld (the metadataPayload as JSON)
    """
    url = req.url
    tables = await parse_excel_or_csv_from_url(url)
        
    # Prepare async http client
    async with httpx.AsyncClient() as client:
        zip_buf = BytesIO()
        # create zip in-memory
        import zipfile
        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            # add a pretty-printed metadata JSON-LD file
            #try:
            #    metadata_bytes = json.dumps(metadata, ensure_ascii=False, indent=2).encode("utf-8")
            #except Exception:
            #    metadata_bytes = str(metadata).encode("utf-8")
            #zf.writestr("metadata.jsonld", metadata_bytes)

            # iterate tables and produce CSV per table
            for idx, tbl in enumerate(tables):
                # table can be a simple string or an object
                # prefer tbl['url'] if available; allow url to be list or string
                tbl_url = None
                if isinstance(tbl, str):
                    tbl_url = tbl
                elif isinstance(tbl, dict):
                    u = tbl.get("url")
                    if isinstance(u, list) and len(u) > 0:
                        tbl_url = u[0]
                    else:
                        tbl_url = u
                # fallback name
                base = f"table_{idx+1}"
                filename_hint = safe_name_from_url(str(tbl_url) if tbl_url else "", base)
                csv_filename = filename_hint
                # ensure extension .csv
                if not csv_filename.lower().endswith(".csv"):
                    csv_filename = csv_filename + ".csv"

                # Attempt to fetch the CSV at tbl_url
                got_text = None
                if tbl_url:
                    got_text = await fetch_text(client, tbl_url)

                if got_text:
                    # if the remote resource looks like CSV already, use it
                    # But if remote is JSON or other, we still include the text as .csv (user can inspect)
                    zf.writestr(csv_filename, got_text.encode("utf-8"))
                else:
                    # no remote data — try to generate CSV header from tableSchema.columns
                    headers = []
                    rows = None
                    if isinstance(tbl, dict):
                        ts = tbl.get("tableSchema") or {}
                        cols = ts.get("columns") if ts else None
                        if isinstance(cols, list):
                            # columns may be objects with titles/name; try to get 'name' or 'titles'
                            hdrs = []
                            for c in cols:
                                if isinstance(c, dict):
                                    # titles can be string or array; handle both
                                    t = c.get("name") or c.get("titles") or c.get("title")
                                    if isinstance(t, list) and len(t) > 0:
                                        hdrs.append(str(t[0]))
                                    elif t is not None:
                                        hdrs.append(str(t))
                                    else:
                                        # fallback: propertyUrl or blank
                                        hdrs.append(c.get("propertyUrl") or "")
                                else:
                                    # if column is a string
                                    hdrs.append(str(c))
                            headers = [h if h is not None else "" for h in hdrs]
                        else:
                            headers = []
                    else:
                        headers = []

                    if not headers:
                        # fallback to single column with table name
                        headers = [filename_hint.replace(".csv", "")]

                    csv_text = build_csv_text_from_rows(headers, rows)
                    zf.writestr(csv_filename, csv_text.encode("utf-8"))

        # finalize zip buffer
        zip_buf.seek(0)
        # StreamingResponse is fine for in-memory buffer
        headers = {
            "Content-Disposition": 'attachment; filename="csvw_tables.zip"'
        }
        return StreamingResponse(zip_buf, media_type="application/zip", headers=headers)

@app.post("/suggest", summary="Suggest a CSVW configuration",
    description="""Suggest a CSVW configuration or a given tableor excel""" )
async def suggest(req: URLRequest):
    """
    build a CSVW metadata (JSON-LD) using tables[] with 
    guessed columns/datatype/primaryKey.
    """
    url = req.url

    tables = await parse_excel_or_csv_from_url(url)
    # Build CSVW tables[] array
    csvw_tables = []
    all_headers = []
    for t in tables:
        headers = t.get("headers", []) or []
        rows = t.get("rows", []) or []
        # sample for types: first up to 5 rows
        sample_for_types = rows[:5]
        # sample for uniqueness for primary key: up to 200 rows
        sample_for_uniqueness = rows[:200]
        # build columns
        columns = []
        for h in headers:
            samples = [r.get(h, "") for r in sample_for_types]
            guessed = guess_type_from_samples(samples)
            datatype = None
            if guessed == "number":
                datatype = "http://www.w3.org/2001/XMLSchema#decimal"
            if guessed == "date":
                datatype = "http://www.w3.org/2001/XMLSchema#dateTime"
            prop = "schema:value"
            # small mapping for common names
            keyLower = h.lower() if isinstance(h, str) else ""
            if "name" == keyLower or keyLower.endswith("name"):
                prop = "schema:name"
            elif "email" in keyLower:
                prop = "schema:email"
            elif keyLower in ("id","identifier","uuid"):
                prop = "schema:identifier"
            columns.append({
                "titles": h,
                "name": h,
                "dcterms:description": h,
                "propertyUrl": prop,
                **({"datatype": datatype} if datatype else {})
            })
        # guess primaryKey
        pk = pick_primary_key(headers, sample_for_uniqueness)
        table = {
            "url": t.get("url"),
            "tableSchema": {
                "columns": columns
            }
        }
        if pk:
            table["tableSchema"]["primaryKey"] = pk
        csvw_tables.append(table)
        for h in headers:
            if h not in all_headers:
                all_headers.append(h)
    # build compact context mapping for readability
    #context_map = {"@vocab": "http://example.org/vocab#"}
    #for h in all_headers:
    #    token = "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in str(h))
    #    context_map[h] = f"http://example.org/vocab#{token}"
    metadata = {
        "@context": ["https://www.w3.org/ns/csvw.jsonld", context_map],
        "tables": csvw_tables
    }
    return {"metadata": metadata, "tables": tables}


@app.post("/convert", summary="Converts one or more tables to a knowledge graph, from a csvw configuration",
    description="""converts one or more tables to a knowledge graph, from a csvw configuration""")
async def convert_to_rdf(req: ConvertRequest = Body(...)):
    """
    Convert workbook or CSV at `url` to RDF using provided CSV-W context (object or URL).
    Returns a string with the chosen serialization.
    """
    data = req.data
    context_input = req.context
    fmt = (req.output_format or "ttl").lower()


    allowed_formats = "ttl,rdfxml,json-ld,sqlite,gpkg".split(',')

    if fmt not in allowed_formats:
        raise HTTPException(status_code=400, detail=f"Output_format must be one of: {', '.join(allowed_formats)}")

    # 0) get context
    if isinstance(context_input, str):
        content = await fetch_bytes(context_input)
        context_input = json.loads(content.decode("utf-8", errors="replace"))

    # 1) if url is none, fetch from context
    if not data or len(data) == 0:
        data = []
        for t in context_input.get('tables',[]):
            if t.get('url'):
                data.append(t.get('url'))

    # 2) convert data to graph
    graph = CSVWConverter.to_rdf(next(iter(data),None), 
                                 context_input, 
                                 mode='minimal')
    
    if isinstance(graph, Graph):
        if fmt == "ttl":
            return Response(content=graph.serialize(format="turtle"), 
                            media_type="text/turtle")
        if fmt == "rdfxml":
            return Response(content=graph.serialize(format="xml"), 
                            media_type="application/rdf+xml")
        if fmt == "json-ld":
            return Response(content=graph.serialize(format="json-ld"), 
                            media_type="application/ld+json")
        # --- BEGIN: Enhanced SQLite export with SOSA mapping ---
        if fmt == "sqlite" or fmt == "gpkg":
            db = rdf2rdb(graph)
            db_bytes = io.BytesIO()
            for line in db.iterdump():  # iterdump gives SQL statements to recreate DB
                db_bytes.write(f"{line}\n".encode())
            db.close()
            db_bytes.seek(0)

            return Response(content=db_bytes.getvalue(), media_type="application/sqlite",
                            headers={"Content-Disposition": "attachment; filename=export.db"})
    else:
        raise HTTPException(status_code=500, detail=f"Parsing error")


# ---- Helpers ----
def bn_to_urn(node):
    """Create a stable URN for a Blank Node or non-URI node."""
    s = str(node)
    h = hashlib.sha1(s.encode("utf-8")).hexdigest()
    return f"urn:bn:{h}"

def node_to_uri(node):
    """Return a stable string identifier for a node (URI or BNode)."""
    if isinstance(node, URIRef):
        return str(node)
    if isinstance(node, BNode):
        return bn_to_urn(node)
    if isinstance(node, Literal):
        return str(node)
    return str(node)

def first_value(g, subject, predicate):
    """Return the first value for subject predicate or None (as python object)."""
    v = next(g.objects(subject, predicate), None)
    if v is None:
        return None
    # For rdflib LITERAL / URIRef return python value appropriately
    if isinstance(v, Literal):
        return str(v)
    return v  # keep URIRef or BNode for further lookup

def label_for(g, node):
    """Prefer rdfs:label, then skos:prefLabel if present, else the URI / bnode string."""
    if node is None:
        return None
    lab = first_value(g, node, RDFS.label) or first_value(g, node, SKOS.prefLabel) or first_value(g, node, DCTERMS.title)
    if lab:
        return str(lab)
    # fallback to node's string/uri
    return node_to_uri(node)

def get_pref_labels_from_remote(concept_uri):
    g = Graph()
    # rdflib will try to GET the resource and parse it (it guesses format)
    try:
        g.parse(concept_uri)  # remove format if server may return another RDF format
        uri = URIRef(concept_uri)
        return list(g.objects(uri, SKOS.prefLabel))
    except Exception as ex:
        print('failed get prefLabel for skos term',concept_uri,ex)

def types_text_for(g, node):
    """
    Return a text representation for rdf:type values on `node`.
    - Return comma-joined localnames of the types (e.g. 'Feature,Profile').
    - Returns None if no types found.
    """
    types = list(g.objects(node, RDF.type))
    if not types:
        return None

    # fallback: collect localnames or full URIs if localname not available
    def localname(uri):
        s = str(uri)
        if "#" in s:
            return s.split("#", 1)[1]
        if "/" in s:
            return s.rsplit("/", 1)[1]
        return s

    names = []
    for t in types:
        if isinstance(t, URIRef):
            names.append(localname(t))
        else:
            names.append(str(t))
    return ",".join(names)

# utility upsert functions returning row id
def upsert_single_return_id(cur, table, uri, label=None, extra_col=None, extra_val=None):
    """
    Generic insert-or-ignore then select id. If extra_col/extra_val provided, include in insert.
    """
    if uri is None:
        return None
    if extra_col:
        sql_ins = f"INSERT OR IGNORE INTO {table} (uri, label, {extra_col}) VALUES (?, ?, ?)"
        cur.execute(sql_ins, (uri, label, extra_val))
    else:
        cur.execute(f"INSERT OR IGNORE INTO {table} (uri, label) VALUES (?, ?)", (uri, label))
    cur.execute(f"SELECT {table}_id FROM {table} WHERE uri = ?", (uri,))
    row = cur.fetchone()
    return row[0] if row else None

def dbinit():
    # Create an in-memory SQLite database
    conn = sqlite3.connect(":memory:")
    cur = conn.cursor()

    # -----------------------------
    # 1. Create tables
    # -----------------------------
    cur.execute("""
    CREATE TABLE result (
        result_uri TEXT PRIMARY KEY,
        value REAL,
        unit_of_measure_id TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE observation (
        observation_uri TEXT PRIMARY KEY,
        result_uri TEXT,
        phenomenon_time TEXT,
        procedure_id TEXT,
        property_id TEXT,
        foi_id TEXT,
        FOREIGN KEY(result_uri) REFERENCES result(result_uri)
    )
    """)

    cur.execute("""
    CREATE TABLE procedure (
        uri TEXT PRIMARY KEY,
        label TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE unitofmeasure (
        uri TEXT PRIMARY KEY,
        label TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE property (
        uri TEXT PRIMARY KEY,
        label TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE feature_of_interest (
        uri TEXT PRIMARY KEY,
        label TEXT
    )
    """)
    return conn

# get sqlite for rdf
def rdf2rdb(g, RDF_FORMAT='turtle'): # "xml", "nt", "json-ld" etc.
    
    conn = dbinit()
    cur = conn.cursor()
    if 1==1: # try:
        # iterate over explicit observation subjects
        obs_count = 0
        for obs in g.subjects(RDF.type, SOSA.Observation):
            obs_count += 1
            obs_uri = node_to_uri(obs)
            # Phenomenon time (literal)
            phen_time = None
            phen_time2 = first_value(g, obs, SOSA.phenomenonTime) or first_value(g, obs, SOSA.phenomenon_time) or first_value(g, obs, SOSA.resultTime)
            if phen_time2:
                try:    
                    dt = datetime.strptime(phen_time2.split('T')[0], "%Y-%m-%d")
                    phen_time = dt.timestamp()
                except:
                    print('can not date parse',phen_time2.split('T')[0])
            # result text (simple literal)
            qual_value_text = first_value(g, obs, SOSA.hasSimpleResult)
            # qualitative value node (may be blank node or URI)
            qual_node = first_value(g, obs, SOSA.result) or first_value(g, obs, SOSA.hasResult)
            qual_uri = None
            qual_label = None
            qual_value_text = None
            if qual_node is not None:
                # if it's a node (URIRef/BNode), create stable id
                qual_uri = node_to_uri(qual_node) if not isinstance(qual_node, Literal) else None
                # label or text
                qual_label = label_for(g, qual_node) if qual_uri else str(qual_node)
                # maybe the observation also has a simple result we should preserve as value_text
                qual_value_text = first_value(g, qual_node, QUDT.numericValue) or first_value(g, qual_node, QUDT.quantityValue) 
                # unit (might be linked via qudt or as property on observation)
                unit_node = first_value(g, qual_node, QUDT.unit) or first_value(g, obs, QUDT.unit)
                unit_uri = node_to_uri(unit_node) if unit_node is not None else None
                unit_label = label_for(g, unit_node) if unit_node is not None else None
            # procedure
            proc_node = first_value(g, obs, SOSA.usedProcedure) or first_value(g, obs, SOSA.isProducedBy)
            proc_uri = node_to_uri(proc_node) if proc_node is not None else None
            proc_label = label_for(g, proc_node) if proc_node is not None else None
            # observed property
            prop_node = first_value(g, obs, SOSA.observedProperty) or first_value(g, obs, SOSA.hasObservedProperty)
            prop_uri = node_to_uri(prop_node) if prop_node is not None else None
            prop_label = label_for(g, prop_node) if prop_node is not None else None
            # feature of interest
            foi_node = first_value(g, obs, SOSA.hasFeatureOfInterest) or first_value(g, obs, SOSA.featureOfInterest)
            foi_uri = node_to_uri(foi_node) if foi_node is not None else None
            foi_label = label_for(g, foi_node) if foi_node is not None else None
            foi_type = types_text_for(g, foi_node)
            # geometry text
            # geom_text = get_geometry_text(g, foi_node) if foi_node is not None else None

            # Upsert referenced entities and get ids
            proc_id = upsert_single_return_id(cur,"procedure", proc_uri, proc_label)
            unit_id = upsert_single_return_id(cur,"unitofmeasure", unit_uri, unit_label)
            prop_id = upsert_single_return_id(cur,"property", prop_uri, prop_label)
            foi_id = upsert_single_return_id(cur,"feature_of_interest", foi_uri, foi_type)
            if qual_value_text not in [None, '']:
                UPSERT_RES = f"INSERT INTO result (result_uri,value,unit_of_measure_id) VALUES (?, ?, ?)"
                cur.execute(UPSERT_RES,(qual_uri,qual_value_text,unit_id))

            UPSERT_OBS = """
            INSERT INTO observation (observation_uri, result_uri, phenomenon_time, procedure_id, property_id, foi_id)
            VALUES (?, ?, ?, ?, ?, ?)
            """
            cur.execute(UPSERT_OBS,(obs_uri, qual_uri, phen_time, proc_id, prop_id, foi_id))
            
        conn.commit()
        print("Data loaded")
    #except Exception as e:
    #    print('error',e)
    #finally: 
        cur.close()
        return conn

    

