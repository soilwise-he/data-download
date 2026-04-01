# ----------------------------------------------------------------------------
#  Copyright (C) 2026 ISRIC - World Soil Information
#  Produced in the Scope of the SoilWISE Project
#  SoilWISE is funded by the European Union’s Horizon Europe research and 
#  innovation programme under grant agreement No 101056973.
# ----------------------------------------------------------------------------

from typing import List, Any, Dict, Optional, Union
import httpx, json, io
import pandas as pd
from io import BytesIO, StringIO
import csv, os
from datetime import datetime


# -----------------------
# Utilities
# -----------------------
async def fetch_bytes(url: str, timeout: int = 30) -> bytes:
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
        r = await client.get(url)
        r.raise_for_status()
        return r.content

def dataframe_to_rows_and_headers(df: pd.DataFrame) -> (List[str], List[Dict[str,Any]]):
    # ensure columns are strings
    headers = [str(c) for c in df.columns.tolist()]
    rows = df.fillna("").to_dict(orient="records")
    return headers, rows

def to_csv_text(headers: List[str], rows: List[Dict[str,Any]]) -> str:
    out = StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h, "") for h in headers])
    return out.getvalue()

# header detection for a sheet represented as list-of-lists
def detect_header_row(raw_rows: List[List[Any]], max_scan: int = 20) -> int:
    def is_empty_row(r):
        if not r:
            return True
        return all((c is None) or (str(c).strip() == "") for c in r)

    max_scan = min(max_scan, len(raw_rows))
    best_idx = 0
    best_score = -1e9

    for i in range(max_scan):
        row = raw_rows[i] or []
        if is_empty_row(row):
            continue
        non_empty = sum(1 for c in row if c is not None and str(c).strip() != "")
        if non_empty == 0:
            continue
        unique_count = len(set(str(c).strip() if c is not None else "" for c in row))
        uniqueness = unique_count / max(1, non_empty)

        # evaluate following rows as "data-like"
        following_good = 0
        following_checked = 0
        for j in range(i+1, min(len(raw_rows), i+1+10)):
            r2 = raw_rows[j] or []
            non_empty2 = sum(1 for c in r2 if c is not None and str(c).strip() != "")
            if non_empty2 == 0:
                following_checked += 1
                continue
            if non_empty2 >= max(1, int(non_empty * 0.5)):
                following_good += 1
            following_checked += 1

        follow_ratio = (following_good / following_checked) if following_checked > 0 else 0.0
        score = non_empty * 0.6 + uniqueness * 2.0 + follow_ratio * 5.0
        if uniqueness < 0.25:
            continue
        if score > best_score:
            best_score = score
            best_idx = i

    # fallback to first non-empty row
    if best_score < -1e8:
        for i in range(min(5, len(raw_rows))):
            if not is_empty_row(raw_rows[i]):
                return i
        return 0
    return best_idx

# guess type from samples (list of values)
def guess_type_from_samples(values: List[Any]) -> str:
    num_count = 0
    date_count = 0
    non_empty = 0
    for v in values:
        s = "" if v is None else str(v).strip()
        if s == "":
            continue
        non_empty += 1
        # numeric?
        try:
            ff = float(s.replace(",", "."))
            if pd.notna(ff):
                num_count += 1
                continue
        except Exception:
            pass
        # date-like?
        try:
            parsed = pd.to_datetime(s, errors="coerce")
            if pd.notna(parsed):
                date_count += 1
                continue
        except Exception:
            pass
    if non_empty == 0:
        return "string"
    if num_count / non_empty >= 0.8:
        return "number"
    if date_count / non_empty >= 0.8:
        return "date"
    return "string"

# pick primary key candidate from rows (list of dicts)
def pick_primary_key(headers: List[str], rows_sample: List[Dict[str,Any]]) -> Optional[str]:
    # compute unique ratio per header
    scores = []
    for h in headers:
        non_empty = 0
        seen = set()
        for r in rows_sample:
            v = r.get(h, "")
            s = "" if v is None else str(v).strip()
            if s == "":
                continue
            non_empty += 1
            seen.add(s)
        if non_empty == 0:
            uniq_ratio = 0.0
        else:
            uniq_ratio = len(seen) / non_empty
        scores.append((h, non_empty, len(seen), uniq_ratio))
    # prefer name hints
    hints = ['id', 'identifier', 'uuid', 'code', 'key']
    strong = [s for s in scores if s[1] >= 3 and s[3] >= 0.98]
    good = [s for s in scores if s[1] >= 3 and s[3] >= 0.8]
    def pick_by_hint(cands):
        for hint in hints:
            for c in cands:
                if hint in c[0].lower():
                    return c[0]
        return None
    if len(strong) == 1:
        return strong[0][0]
    if len(strong) > 1:
        return pick_by_hint(strong) or sorted(strong, key=lambda x: -x[3])[0][0]
    if good:
        return pick_by_hint(good) or sorted(good, key=lambda x: -x[3])[0][0]
    # final attempt: perfect uniqueness
    perfect = [s for s in scores if s[1] >= 3 and s[2] == s[1]]
    if perfect:
        return pick_by_hint(perfect) or perfect[0][0]
    return None

# -----------------------
# Core parsing logic
# -----------------------
async def parse_excel_or_csv_from_url(url: str) -> List[Dict]:
    """
    Returns a list of table entries:
      [{ "url": "<sheet-identifier-or-url>", "sheet_name": "<...>", "headers": [...], "rows": [...], "csv_text": "..." }, ...]
    """
    content = await fetch_bytes(url)
    # try to detect if excel by signature (PK header for xlsx) or by extension
    lower = url.lower()
    is_excel_ext = lower.endswith(".xlsx") or lower.endswith(".xls") or lower.endswith(".xlsm") or lower.endswith(".ods")
    # quick signature check: xlsx files are zip (PK\x03\x04)
    if (not is_excel_ext) and content[:4] != b'PK\x03\x04':
        # treat as CSV text
        text = content.decode("utf-8", errors="replace")
        parsed = list(csv.DictReader(StringIO(text)))
        headers = parsed[0].keys() if parsed else []
        headers = [str(h) for h in headers]
        rows = parsed
        csv_text = text
        return [{
            "url": url,
            "sheet_name": None,
            "headers": headers,
            "rows": rows,
            "csv_text": csv_text
        }]
    # else parse as excel using pandas
    bio = BytesIO(content)
    try:
        # sheet_name=None reads all sheets into dict
        xls = pd.read_excel(bio, sheet_name=None, engine="openpyxl")
    except Exception as e:
        # fallback: try using engine 'xlrd' for old xls (if available)
        raise HTTPException(status_code=500, detail=f"Failed to read Excel workbook: {e}")

    results = []
    # pandas returns dict of DataFrame
    for sheet_name, df in xls.items():
        # convert dataframe to raw rows (list-of-lists) to run header detection
        raw_rows = df.fillna(value=pd.NA).astype(object).values.tolist()
        # raw_rows may include header row as first row (pandas by default already uses first row as header)
        # But since we asked sheet_name=None with default behavior, pandas used first non-empty row as header.
        # To be conservative, convert sheet to array-of-arrays using openpyxl if you want raw cell values.
        # Here we use a fallback: if all column names are "Unnamed..." then detect header row from values
        col_names = df.columns.tolist()
        if all(str(c).startswith("Unnamed") for c in col_names):
            # try building raw_rows from values (pandas placed header as Unnamed)
            # Use df.to_records? We'll re-read via openpyxl for better raw data.
            try:
                from openpyxl import load_workbook
                bio.seek(0)
                wb = load_workbook(filename=bio, read_only=True, data_only=True)
                ws = wb[sheet_name]
                raw = []
                for row in ws.iter_rows(values_only=True):
                    raw.append(list(row))
                header_idx = detect_header_row(raw)
                header_row = ["" if c is None else str(c).strip() for c in raw[header_idx]]
                # build rows
                data_rows = []
                empty_streak = 0
                for r in raw[header_idx+1:]:
                    if r is None or all((c is None or str(c).strip()=="") for c in r):
                        empty_streak += 1
                        if empty_streak >= 5:
                            break
                        continue
                    empty_streak = 0
                    rowobj = {}
                    for j, h in enumerate(header_row):
                        key = h or f"col{j+1}"
                        val = r[j] if j < len(r) else ""
                        rowobj[key] = "" if val is None else val
                    data_rows.append(rowobj)
                headers = header_row
                csv_text = to_csv_text(headers, data_rows)
                results.append({
                    "url": f"{url}#{sheet_name}",
                    "sheet_name": sheet_name,
                    "headers": headers,
                    "rows": data_rows,
                    "csv_text": csv_text
                })
                continue
            except Exception:
                # fallback to using pandas parsing below
                pass

        # If we get here, use DataFrame as parsed by pandas.
        headers, rows = dataframe_to_rows_and_headers(df)
        csv_text = to_csv_text(headers, rows)
        results.append({
            "url": f"{url}#{sheet_name}",
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": rows,
            "csv_text": csv_text
        })
    return results

# Helper: safe filename from URL or fallback
def safe_name_from_url(url: str, fallback: str) -> str:
    try:
        parsed = urlparse(url)
        name = parsed.path.split("/")[-1] or parsed.netloc or fallback
        # remove fragments and query parts if present
        name = name.split("#")[0].split("?")[0]
        if not name:
            return fallback
        return name
    except Exception:
        return fallback

# Helper: build CSV text from header list + (optional) rows (rows as list of dicts)
def build_csv_text_from_rows(headers: List[str], rows: Optional[List[Dict[str, Any]]] = None) -> str:
    output = BytesIO()
    # write text through TextIOWrapper to support newline handling in zip
    text_wr = TextIOWrapper(output, encoding="utf-8", newline="")
    writer = csv.writer(text_wr, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    if rows:
        for r in rows:
            row = [ (r.get(h) if isinstance(r, dict) else '') for h in headers ]
            writer.writerow(row)
    text_wr.flush()
    result = output.getvalue().decode("utf-8")
    output.close()
    return result

# Async fetch helper with timeout and basic heuristics
async def fetch_text(client: httpx.AsyncClient, url: str, timeout: float = 10.0) -> Optional[str]:
    try:
        resp = await client.get(url, timeout=timeout)
        resp.raise_for_status()
        # try to decode as text
        ct = resp.headers.get("content-type", "")
        # treat as text if content-type mentions text or csv or json
        if "text" in ct or "csv" in ct or "json" in ct or ct == "" or "application/octet-stream" in ct:
            # prefer resp.text which decodes according to charset
            return resp.text
        # otherwise try to decode with utf-8 fallback
        return resp.content.decode("utf-8", errors="replace")
    except Exception:
        return None